from django.core.exceptions import ValidationError
from openpyxl.styles import Alignment
from django.shortcuts import render
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from .models import Proveedor
from .forms import ProveedorForm
from django.shortcuts import render, redirect
from .models import Proveedor, Producto, Departamento, Empleado, Almacen, EntradaAlmacen, Pedido
from django.contrib import messages
from .forms import ProductoForm, ProveedorForm, DepartamentoForm, EmpleadoForm, AlmacenForm, EntradaAlmacenForm, PedidoForm, BuscarForm, BuscarProveedoresForm, BuscarProductosForm
from django.shortcuts import render, get_object_or_404
from django.contrib.auth import authenticate, login
from django.views import View
from django.db.models import Q
from openpyxl import Workbook
from django.http import JsonResponse
import matplotlib.pyplot as plt
from django.db.models import Sum, F
from django.db.models.functions import ExtractMonth
import calendar
from datetime import datetime
from decimal import Decimal
from django.http import HttpResponse
import json

# Create your views here.


def inicio(request):

    return render(request, "inicio.html")


def registrarProveedor(request):
    buscar_form = BuscarProveedoresForm()
    registro_form = ProveedorForm()
    proveedores = Proveedor.objects.all()

    if request.method == 'POST':
        buscar = request.POST.get('buscar')

        if buscar:
            proveedores = proveedores.filter(
                Q(nombre__icontains=buscar) |
                Q(idProveedor__icontains=buscar)).distinct()

        else:
            # Procesar el formulario de registro
            registro_form = ProveedorForm(request.POST)
            if registro_form.is_valid():
                registro_form.save()
                messages.success(request, 'Registro Exitoso!')
                return redirect('registrar_proveedor')

    # Manejar el escaneo del código de barras
    if request.method == 'POST' and 'codigo_barras' in request.POST:
        codigo_barras = request.POST.get('codigo_barras')
        registro_form = ProveedorForm(initial={'sku': codigo_barras})

    context = {
        'buscar_form': buscar_form,
        'registro_form': registro_form,
        'proveedores': proveedores
    }

    return render(request, 'registroProveedor.html', context)


def edicionProveedor(request, codigo):
    proveedores = get_object_or_404(Proveedor, idProveedor=codigo)
    if request.method == 'POST':
        form = ProveedorForm(request.POST, instance=proveedores)
        if form.is_valid():
            form.save()
            messages.success(request, 'Proveedor actualizado correctamente.')
            return redirect('registrar_proveedor')
    else:
        form = ProveedorForm(instance=proveedores)
    return render(request, 'ActualizarProveedor.html', {'form': form, 'proveedores': proveedores})


def actualizar_proveedor(request, codigo):
    proveedores = get_object_or_404(Proveedor, idProveedor=codigo)

    if request.method == 'POST':
        form = ProveedorForm(request.POST, instance=proveedores)
        if form.is_valid():
            form.save()
            messages.success(request, 'Proveedor actualizado correctamente.')
            # Redirecciona a la página de inicio o a la página relevante después de guardar
            return redirect('registrar_proveedor')
    else:
        form = ProveedorForm(instance=proveedores)

    return render(request, 'ActualizarProveedor.html', {'form': form, 'proveedores': proveedores})


def eliminarProveedor(request, codigo):
    proveedores = Proveedor.objects.get(idProveedor=codigo)
    proveedores.delete()
    messages.success(request, 'Eliminado!')
    return redirect('registrar_proveedor')


def registrarProducto(request):
    buscar_form = BuscarProductosForm()
    registro_form = ProductoForm()
    proveedores = Proveedor.objects.all()
    productos = Producto.objects.all()

    if request.method == 'POST':
        buscar = request.POST.get('buscar')

        if buscar:
            productos = productos.filter(
                Q(nombre__icontains=buscar) |
                Q(sku__icontains=buscar)).distinct()

        else:
            # Procesar el formulario de registro
            registro_form = ProductoForm(request.POST)
            if registro_form.is_valid():
                sku = registro_form.cleaned_data['sku']
                existencia_producto = Producto.objects.filter(sku=sku)
                if existencia_producto.exists():
                    messages.error(request, 'SKU duplicado')
                else:
                    registro_form.save()
                    messages.success(request, 'Registro Exitoso!')
                    return redirect('registrar_producto')

    context = {
        'buscar_form': buscar_form,
        'registro_form': registro_form,
        'proveedores': proveedores,
        'productos': productos
    }

    return render(request, 'registroProducto.html', context)


def edicionProducto(request, codigo):
    producto = get_object_or_404(Producto, sku=codigo)

    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            form.save()
            messages.success(request, 'Producto actualizado exitosamente!')
            return redirect('registrar_producto')
    else:
        form = ProductoForm(instance=producto)

    return render(request, 'actualizarProducto.html', {'form': form, 'producto': producto})


def eliminarProducto(request, codigo):
    producto = Producto.objects.get(sku=codigo)
    producto.delete()
    messages.success(request, 'Eliminado!')
    return redirect('registrar_producto')


def actualizarProducto(request, codigo):
    producto = get_object_or_404(Producto, sku=codigo)
    form = ProductoForm(instance=producto)

    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            form.save()
            messages.success(request, 'Producto actualizado exitosamente!')
            return redirect('registrar_producto')

    return render(request, 'actualizarProducto.html', {'form': form, 'producto': producto})


def registrarDepartamento(request):
    form = DepartamentoForm()
    departamentos = Departamento.objects.all

    if request.method == 'POST':

        form = DepartamentoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Registro Exitoso!')
            return redirect('registroDepartamento')

    return render(request, 'registroDepartamento.html', {'form': form, 'departamentos': departamentos})


def edicionDepartamento(request, codigo):
    departamentos = get_object_or_404(Departamento, idDepartamento=codigo)
    if request.method == 'POST':
        form = DepartamentoForm(request.POST, instance=departamentos)
        if form.is_valid():
            form.save()
            messages.success(
                request, 'Departamento actualizado correctamente.')
            return redirect('registroDepartamento')
    else:
        form = DepartamentoForm(instance=departamentos)
    return render(request, 'actualizarDepartamento.html', {'form': form, 'departamentos': departamentos})


def actualizarDepartamento(request, codigo):
    departamentos = get_object_or_404(Departamento, idDepartamento=codigo)
    form = DepartamentoForm(instance=departamentos)

    if request.method == 'POST':
        form = DepartamentoForm(request.POST, instance=departamentos)
        if form.is_valid():
            form.save()
            messages.success(request, 'Departamento actualizado exitosamente!')
            return redirect('registroDepartamento')

    return render(request, 'actualizarDepartamento.html', {'form': form, 'departamentos': departamentos})


def eliminarDepartamento(request, codigo):
    departamentos = Departamento.objects.get(idDepartamento=codigo)
    departamentos.delete()
    messages.success(request, 'Eliminado!')
    return redirect('registroDepartamento')


def registrarEmpleado(request):
    buscar_form = BuscarForm()  # Formulario para buscar empleados
    registro_form = EmpleadoForm()  # Formulario para registrar un nuevo empleado
    departamentos = Departamento.objects.all()
    empleados = Empleado.objects.all()

    if request.method == 'POST':
        buscar = request.POST.get('buscar')

        if buscar:
            empleados = empleados.filter(
                Q(nombre__icontains=buscar) |
                Q(idEmpleado__icontains=buscar)).distinct()
        else:
            # Procesar el formulario de registro
            registro_form = EmpleadoForm(request.POST)
            if registro_form.is_valid():
                registro_form.save()
                messages.success(request, 'Registro Exitoso!')
                return redirect('registroEmpleado')

    context = {
        'buscar_form': buscar_form,
        'registro_form': registro_form,
        'empleados': empleados,
        'departamentos': departamentos
    }
    return render(request, 'registroEmpleado.html', context)


def actualizarEmpleado(request, codigo):
    empleados = get_object_or_404(Empleado, idEmpleado=codigo)
    form = EmpleadoForm(instance=empleados)

    if request.method == 'POST':
        form = EmpleadoForm(request.POST, instance=empleados)
        if form.is_valid():
            form.save()
            messages.success(request, 'Empleado actualizado exitosamente!')
            return redirect('registroEmpleado')

    return render(request, 'actualizarEmpleado.html', {'form': form, 'empleados': empleados})


def edicionEmpleado(request, codigo):
    empleados = get_object_or_404(Empleado, idEmpleado=codigo)

    if request.method == 'POST':
        form = EmpleadoForm(request.POST, instance=empleados)
        if form.is_valid():
            form.save()
            messages.success(request, 'Empleado actualizado exitosamente!')
            return redirect('registroEmpleado')
    else:
        form = EmpleadoForm(instance=empleados)

    return render(request, 'actualizarEmpleado.html', {'form': form, 'empleados': empleados})


def eliminarEmpleado(request, codigo):
    empleados = Empleado.objects.get(idEmpleado=codigo)
    empleados.delete()
    messages.success(request, 'Eliminado!')
    return redirect('registroEmpleado')


def registrarAlmacen(request):
    form = AlmacenForm()
    almacenes = Almacen.objects.all()

    if request.method == 'POST':
        buscar = request.POST.get('buscar')

        if buscar:
            almacenes = almacenes.filter(
                Q(idAlmacen__icontains=buscar) |
                Q(Fksku_id__nombre__icontains=buscar) |
                Q(Fksku_id__sku__icontains=buscar)).distinct()

    if request.method == 'POST':
        form = AlmacenForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Registro Exitoso!')
            return redirect('registrar_almacen')

    if request.method == 'POST':
        form = EntradaAlmacenForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Registro Exitoso!')
            return redirect('registrar_EntradaAlmacen')

    # ESTE ES MI CODIGO PARA LLENAR LOS DATOS AUTOMATICAMENTE APARTIR DEL CODIGO DE BARRAS
    if request.method == 'POST':
        codigo_barras = request.POST.get('codigoBarras')
        try:
            producto = Producto.objects.get(codigoBarras=codigo_barras)
            form = AlmacenForm(initial={
                'codigoBarras': codigo_barras,
                'nombreFksu': producto.nombre,
                'Fksku': producto.sku
            })
        except Producto.DoesNotExist:
            form = AlmacenForm()
            # Asignar el error manualmente al atributo errors
            form.errors['codigoBarras'] = ['El producto no existe']
    else:
        form = AlmacenForm()

    paginator = Paginator(almacenes, 20)
    page = request.GET.get('page')  # Obtener el número de página actual

    try:
        almacenes = paginator.page(page)
    except PageNotAnInteger:
        almacenes = paginator.page(1)
    except EmptyPage:
        almacenes = paginator.page(paginator.num_pages)

    context = {

        'form': form,
        'almacenes': almacenes
    }

    return render(request, 'registroAlmacen.html', context)


def edicionAlmacen(request, codigo):
    almacenes = get_object_or_404(Almacen, idAlmacen=codigo)

    if request.method == 'POST':
        form = AlmacenForm(request.POST, instance=almacenes)
        if form.is_valid():
            form.save()
            messages.success(request, 'Almacen actualizado exitosamente!')
            return redirect('registro_almacen')
    else:
        form = AlmacenForm(instance=almacenes)

    return render(request, 'actualizarAlmacen.html', {'form': form, 'almacenes': almacenes})


def actualizarAlmacen(request, codigo):
    almacen = get_object_or_404(Almacen, idAlmacen=codigo)

    if request.method == 'POST':
        cantidad = request.POST.get('cantidad')
        almacen.cantidad = cantidad
        almacen.save()
        messages.success(request, 'Almacen actualizado exitosamente!')
        return redirect('registrar_almacen')


def eliminarAlmacen(request, codigo):
    almacenes = Almacen.objects.get(idAlmacen=codigo)
    almacenes.delete()
    messages.success(request, 'Eliminado!')
    return redirect('registrar_almacen')


def registrarEntradaAlmacen(request):
    form = EntradaAlmacenForm()
    entradaAlmacenes = EntradaAlmacen.objects.all()

    # Filtrar entradas por búsqueda
    if request.method == 'POST':
        buscar = request.POST.get('buscar')

        if buscar:
            entradaAlmacenes = entradaAlmacenes.filter(
                Q(idEntradaAlmacen__icontains=buscar) |
                Q(Fksku_id__nombre__icontains=buscar) |
                Q(Fksku_id__sku__icontains=buscar)).distinct()

    # Procesar el formulario de registro de entrada
    if request.method == 'POST':
        form = EntradaAlmacenForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Registro Exitoso!')
            return redirect('registrar_EntradaAlmacen')

    # Procesar el escaneo de código de barras
    if request.method == 'POST':
        codigo_barras = request.POST.get('codigoBarras')
        try:
            producto = Producto.objects.get(codigoBarras=codigo_barras)
            form = EntradaAlmacenForm(initial={
                'codigoBarras': codigo_barras,
                'nombreFksu': producto.nombre,
                'Fksku': producto.sku
            })
        except Producto.DoesNotExist:
            form = EntradaAlmacenForm()
            form.errors['codigoBarras'] = ['El producto no existe']
    else:
        form = EntradaAlmacenForm()

    # Paginación de la tabla de entradas
    # Mostrar 10 elementos por página
    paginator = Paginator(entradaAlmacenes, 8)
    page = request.GET.get('page')  # Obtener el número de página actual

    try:
        entradaAlmacenes_paginadas = paginator.page(page)
    except PageNotAnInteger:
        entradaAlmacenes_paginadas = paginator.page(1)
    except EmptyPage:
        entradaAlmacenes_paginadas = paginator.page(paginator.num_pages)

    context = {
        'form': form,
        'entradaAlmacenes': entradaAlmacenes_paginadas
    }

    return render(request, 'registroEntradaAlmacen.html', context)


def edicionEntradaAlmacen(request, codigo):
    entradaAlmacenes = get_object_or_404(
        EntradaAlmacen, idEntradaAlmacen=codigo)

    if request.method == 'POST':
        form = EntradaAlmacenForm(request.POST, instance=entradaAlmacenes)
        if form.is_valid():
            form.save()
            messages.success(request, 'Almacen actualizado exitosamente!')
            return redirect('registrar_EntradaAlmacen')
    else:
        form = EntradaAlmacenForm(instance=entradaAlmacenes)

    return render(request, 'actualizarEntradaAlmacen.html', {'form': form, 'entradaAlmacenes': entradaAlmacenes})


def actualizarEntradaAlmacen(request, codigo):
    entradaAlmacen = get_object_or_404(
        EntradaAlmacen, idEntradaAlmacen=codigo)

    if request.method == 'POST':
        fechaEntrada = request.POST.get('fechaEntrada')
        cantidad = request.POST.get('cantidad')
        entradaAlmacen.cantidad = cantidad
        entradaAlmacen.fechaEntrada = fechaEntrada
        entradaAlmacen.save()
        messages.success(
            request, 'Entrada de Almacen actualizado exitosamente!')
        return redirect('registrar_EntradaAlmacen')


def eliminarEntradaAlmacen(request, codigo):
    entradaAlmacenes = EntradaAlmacen.objects.get(idEntradaAlmacen=codigo)
    entradaAlmacenes.delete()
    messages.success(request, 'Eliminado!')
    return redirect('registrar_EntradaAlmacen')


def registrarPedido(request):
    form = PedidoForm()
    pedidos = Pedido.objects.all()
    response_data = {}  # Datos de respuesta para JSON

    if request.method == 'POST':
        form = PedidoForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                response_data['success'] = True
                response_data['message'] = 'Registro Exitoso!'
            except ValidationError as e:
                response_data['success'] = False
                response_data['message'] = str(e)
        else:
            errors = []
            for field, error_list in form.errors.items():
                for error in error_list:
                    errors.append(f"{field}: {error}")
            response_data['success'] = False
            response_data['errors'] = errors

        return JsonResponse(response_data)

    if request.method == 'GET':
        codigo_barras = request.GET.get('codigoBarras')
        if codigo_barras:
            try:
                producto = Producto.objects.get(codigoBarras=codigo_barras)
                form = PedidoForm(initial={
                    'codigoBarras': codigo_barras,
                    'nombreFksu': producto.nombre,
                    'Fksku': producto.sku
                })
            except Producto.DoesNotExist:
                form = PedidoForm()
                response_data['success'] = False
                response_data['message'] = 'El producto no existe.'

    return render(request, 'registroPedido.html', {'form': form, 'pedidos': pedidos})


def edicionPedido(request, codigo):
    pedidos = get_object_or_404(
        Pedido, idPedido=codigo)

    if request.method == 'POST':
        form = PedidoForm(request.POST, instance=pedidos)
        if form.is_valid():
            form.save()
            messages.success(request, 'Almacen actualizado exitosamente!')
            return redirect('registrar_Pedido')
    else:
        form = PedidoForm(instance=pedidos)

    return render(request, 'actualizarPedido.html', {'form': form, 'pedidos': pedidos})


def actualizarPedido(request, codigo):
    pedidos = get_object_or_404(
        Pedido, idPedido=codigo)

    if request.method == 'POST':
        fechaPedido = request.POST.get('fechaPedido')
        cantidad = request.POST.get('cantidad')
        pedidos.cantidad = cantidad
        pedidos.fechaPedido = fechaPedido
        pedidos.save()
        messages.success(
            request, 'Entrada de Almacen actualizado exitosamente!')
        return redirect('registrar_Pedido')


def eliminarPedido(request, codigo):
    pedidos = Pedido.objects.get(idPedido=codigo)
    pedidos.delete()
    messages.success(request, 'Eliminado!')
    return redirect('registrar_Pedido')


def reportePedido(request):
    pedidos = Pedido.objects.all()
    return render(request, "reportePedido.html", {'pedidos': pedidos})


def reporteIngresos(request):
    entradaAlmacenes = EntradaAlmacen.objects.all()
    return render(request, "reporteIngresos.html", {'entradaAlmacenes': entradaAlmacenes})


class GenerarReportePedidosView(View):
    def post(self, request):
        context = {}
        fecha_inicio = request.POST.get(
            'fecha_inicio') if request.POST.get('fecha_inicio') else ''
        fecha_fin = request.POST.get(
            'fecha_fin') if request.POST.get('fecha_fin') else ''
        nombre_producto = request.POST.get(
            'nombre_producto') if request.POST.get('nombre_producto') else ''
        fksku = request.POST.get('fksku') if request.POST.get('fksku') else ''
        nombre_empleado = request.POST.get(
            'nombre_empleado') if request.POST.get('nombre_empleado') else ''
        fksku = request.POST.get(
            'idEmpleado') if request.POST.get('idEmpleado') else ''

        context['fecha_inicio'] = fecha_inicio
        context['fecha_fin'] = fecha_fin
        context['nombre_producto'] = nombre_producto
        context['nombre_empleado'] = nombre_empleado
        context['fksku'] = fksku

        pedidos = Pedido.objects.filter(
            fechaPedido__range=[fecha_inicio, fecha_fin],
            Fksku__nombre__icontains=nombre_producto,
            Fksku__sku__icontains=fksku,
            idEmpleado__idEmpleado__icontains=nombre_empleado
        )
        context['pedidos'] = pedidos
        return render(request, 'reportePedido.html', context)


class GenerarReportePedidosGraf(View):
    def post(self, request):
        context = {}
        fecha_inicio = request.POST.get('fecha_inicio', '')
        fecha_fin = request.POST.get('fecha_fin', '')
        nombre_producto = request.POST.get('nombre_producto', '')
        fksku = request.POST.get('fksku', '')

        context['fecha_inicio'] = fecha_inicio
        context['fecha_fin'] = fecha_fin
        context['nombre_producto'] = nombre_producto
        context['fksku'] = fksku

        pedidos = Pedido.objects.filter(
            fechaPedido__range=[fecha_inicio, fecha_fin],
            Fksku__nombre__icontains=nombre_producto,
            Fksku__sku__icontains=fksku
        ).annotate(month=ExtractMonth('fechaPedido')).values('month').annotate(total=Sum(F('cantidad') * F('Fksku__precio'))).order_by('month')

        valores_pedidos = [float(pedido['total']) for pedido in pedidos]
        context['valores_pedidos'] = valores_pedidos

        return render(request, "graficoPedido.html", context)


class ReporteExcelPedidos(View):
    def get(self, request):
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        nombre_empleado = request.GET.get('nombre_empleado')
        nombre_producto = request.GET.get('nombre_producto')
        fksku = request.GET.get('fksku')

        pedidos = Pedido.objects.filter(
            Q(fechaPedido__range=[fecha_inicio, fecha_fin]),
            Q(Fksku__sku__icontains=fksku),
            Q(idEmpleado__idEmpleado__icontains=nombre_empleado),
            Q(Fksku__nombre__icontains=nombre_producto))

        wb = Workbook()
        ws = wb.active
        ws['B1'] = "REPORTE DE PEDIDOS"

        ws.merge_cells("B1:E1")
        ws['B3'] = 'ID Pedido'
        ws['C3'] = 'Nombre del Empleado'
        ws['D3'] = 'Nombre del Producto'
        ws['E3'] = 'FK Sku'
        ws['F3'] = 'Fecha de Pedido'
        ws['G3'] = 'Cantidad'
        ws['H3'] = 'Precio Unitario'
        ws['I3'] = 'Valor Total'

        cont = 4
        for x in pedidos:
            ws.cell(row=cont, column=2).value = x.idPedido
            ws.cell(row=cont,  column=3).value = x.idEmpleado.idEmpleado
            ws.cell(row=cont, column=4).value = x.Fksku.nombre
            ws.cell(row=cont, column=5).value = x.Fksku.sku
            ws.cell(row=cont, column=6).value = x.fechaPedido
            ws.cell(row=cont, column=7).value = x.cantidad
            ws.cell(row=cont, column=8).value = x.Fksku.precio
            ws.cell(row=cont, column=9).value = x.cantidad * x.Fksku.precio
            cont += 1

        # Establecer el ancho en píxeles para las columnas
        # Anchuras deseadas en píxeles
        column_widths = [22, 22, 37, 22, 22, 22, 22, 22]
        for i, width in enumerate(column_widths, start=2):
            # Convertir el índice de columna a letra
            column_letter = chr(64 + i)
            ws.column_dimensions[column_letter].width = width

        # Aplicar formato y centrar el contenido de todas las celdas
        align = Alignment(horizontal='center', vertical='center')
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = align

        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename=Reporte-Pedidos.xlsx"
        response['Content-Disposition'] = content
        wb.save(response)
        return response


class ReporteExcelAlmacen(View):
    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws['B1'] = "REPORTE DE ALMACEN"

        ws.merge_cells("B1:F1")
        ws['B3'] = 'ID Almacen'
        ws['C3'] = 'Nombre del Producto'
        ws['D3'] = 'FK SKU'
        ws['E3'] = 'Codigo de Barras'
        ws['F3'] = 'Cantidad'

        cont = 4
        almacenes = Almacen.objects.all()
        for x in almacenes:
            ws.cell(row=cont, column=2).value = x.idAlmacen
            ws.cell(row=cont, column=3).value = x.Fksku.nombre
            ws.cell(row=cont, column=4).value = x.Fksku.sku
            ws.cell(row=cont, column=5).value = x.Fksku.codigoBarras
            ws.cell(row=cont, column=6).value = x.cantidad
            cont += 1

        # Establecer el ancho en píxeles para las columnas
        # Anchuras deseadas en píxeles
        column_widths = [38, 38, 38, 38, 38, 38]
        for i, width in enumerate(column_widths, start=2):
            # Convertir el índice de columna a letra
            column_letter = chr(64 + i)
            ws.column_dimensions[column_letter].width = width

        # Aplicar formato y centrar el contenido de todas las celdas
        align = Alignment(horizontal='center', vertical='center')
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = align

        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename=Reporte-Almacen.xlsx"
        response['Content-Disposition'] = content
        wb.save(response)
        return response


class GenerarReporteIngresoView(View):
    def post(self, request):
        context = {}
        fecha_inicio = request.POST.get(
            'fecha_inicio') if request.POST.get('fecha_inicio') else ''
        fecha_fin = request.POST.get(
            'fecha_fin') if request.POST.get('fecha_fin') else ''
        nombre_producto = request.POST.get(
            'nombre_producto') if request.POST.get('nombre_producto') else ''
        fksku = request.POST.get('fksku') if request.POST.get('fksku') else ''

        context['fecha_inicio'] = fecha_inicio
        context['fecha_fin'] = fecha_fin
        context['nombre_producto'] = nombre_producto
        context['fksku'] = fksku

        entradaAlmacenes = EntradaAlmacen.objects.filter(
            fechaEntrada__range=[fecha_inicio, fecha_fin],
            Fksku__nombre__icontains=nombre_producto,
            Fksku__sku__icontains=fksku
        )
        context['entradaAlmacenes'] = entradaAlmacenes
        return render(request, 'reporteIngresos.html', context)


class ReporteExcel(View):
    def get(self, request):
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')

        nombre_producto = request.GET.get('nombre_producto')
        fksku = request.GET.get('fksku')

        entradaAlmacenes = EntradaAlmacen.objects.filter(
            Q(fechaEntrada__range=[fecha_inicio, fecha_fin]),
            Q(Fksku__sku__icontains=fksku),
            Q(Fksku__nombre__icontains=nombre_producto))

        wb = Workbook()
        ws = wb.active
        ws['B1'] = "REPORTE DE INGRESOS"

        ws.merge_cells("B1:E1")
        ws['B3'] = 'ID Entrada'
        ws['C3'] = 'Nombre del Producto'
        ws['D3'] = 'FK Sku'
        ws['E3'] = 'Fecha de Ingreso'
        ws['F3'] = 'Cantidad'
        ws['G3'] = 'Precio Unitario'
        ws['H3'] = 'Valor Total'

        cont = 4
        for x in entradaAlmacenes:
            ws.cell(row=cont, column=2).value = x.idEntradaAlmacen
            ws.cell(row=cont,  column=3).value = x.Fksku.nombre
            ws.cell(row=cont, column=4).value = x.Fksku.sku
            ws.cell(row=cont, column=5).value = x.fechaEntrada
            ws.cell(row=cont, column=6).value = x.cantidad
            ws.cell(row=cont, column=7).value = x.Fksku.precio
            ws.cell(row=cont, column=8).value = x.cantidad * x.Fksku.precio
            cont += 1

        # Establecer el ancho en píxeles para las columnas
        # Anchuras deseadas en píxeles
        column_widths = [22, 37, 22, 22, 22, 22, 22, 22]
        for i, width in enumerate(column_widths, start=2):
            # Convertir el índice de columna a letra
            column_letter = chr(64 + i)
            ws.column_dimensions[column_letter].width = width

        # Aplicar formato y centrar el contenido de todas las celdas
        align = Alignment(horizontal='center', vertical='center')
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = align

        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename=Reporte-Ingresos.xlsx"
        response['Content-Disposition'] = content
        wb.save(response)
        return response


def almacen_view(request):
    if request.method == 'POST':
        codigo_barras = request.POST.get('codigoBarras')
        try:
            producto = Producto.objects.get(codigoBarras=codigo_barras)
            form = EntradaAlmacenForm(initial={
                'codigoBarras': codigo_barras,
                'nombreFksu': producto.nombre,
                'Fksku': producto.sku
            })
        except Producto.DoesNotExist:
            form = EntradaAlmacenForm()
            # Asignar el error manualmente al atributo errors
            form.errors['codigoBarras'] = ['El producto no existe']
    else:
        form = EntradaAlmacenForm()

    context = {'form': form}
    return render(request, 'registroAlmacen.html', context)


def graficoPedido(request):
    year_actual = datetime.now().year
    pedidos_por_mes = []

    for month in range(1, 13):
        # Filtrar pedidos por mes y año
        pedidos_mes = Pedido.objects.filter(
            fechaPedido__year=year_actual,
            fechaPedido__month=month
        )

        # Calcular la sumatoria de los pedidos del mes
        total_mes = sum(float(pedido.cantidad * pedido.Fksku.precio)
                        for pedido in pedidos_mes)

        # Agregar el mes y el total a la lista
        pedidos_por_mes.append((calendar.month_name[month], total_mes))

    # Obtener las listas separadas de meses y valores totales
    meses = [mes for mes, _ in pedidos_por_mes]
    valores_pedidos = [total for _, total in pedidos_por_mes]
    valores_pedidos_json = json.dumps(valores_pedidos)

    context = {'meses': meses, 'valores_pedidos_json': valores_pedidos_json}
    return render(request, "graficoPedido.html", context)


def graficoIngreso(request):

    year_actual = datetime.now().year
    Ingresos_por_mes = []

    for month in range(1, 13):
        # Filtrar pedidos por mes y año
        Ingresos_mes = EntradaAlmacen.objects.filter(
            fechaEntrada__year=year_actual,
            fechaEntrada__month=month
        )

        # Calcular la sumatoria de los pedidos del mes
        total_mes = sum(float(EntradaAlmacen.cantidad * EntradaAlmacen.Fksku.precio)
                        for EntradaAlmacen in Ingresos_mes)

        # Agregar el mes y el total a la lista
        Ingresos_por_mes.append((calendar.month_name[month], total_mes))

    # Obtener las listas separadas de meses y valores totales
    meses = [mes for mes, _ in Ingresos_por_mes]
    valores_ingresos = [total for _, total in Ingresos_por_mes]
    valores_ingresos_json = json.dumps(valores_ingresos)

    context = {'meses': meses, 'valores_ingresos_json': valores_ingresos_json}
    return render(request, "graficoIngreso.html", context)


def graficoPastelp(request):
    departamentos = Departamento.objects.all()
    gastos_por_departamento = []

    for departamento in departamentos:
        # Filtrar pedidos por departamento
        pedidos_departamento = Pedido.objects.filter(
            idEmpleado__fkdepartamento=departamento
        )

        # Calcular el gasto total del departamento
        total_gasto = sum(float(pedido.cantidad * pedido.Fksku.precio)
                          for pedido in pedidos_departamento)

        # Agregar el departamento y el gasto a la lista
        gastos_por_departamento.append((departamento.nombre, total_gasto))

    # Obtener las listas separadas de departamentos y gastos totales
    nombres_departamentos = [
        departamento for departamento, _ in gastos_por_departamento]
    gastos_departamentos = [gasto for _, gasto in gastos_por_departamento]
    gastos_departamentos_json = json.dumps(gastos_departamentos)

    context = {'nombres_departamentos': nombres_departamentos,
               'gastos_departamentos_json': gastos_departamentos_json}
    return render(request, "graficoPastelp.html", context)
